"""Читання офіційного WHO SimpleTabulation ICD-11 MMS Excel-файлу
у ICD11Node (S07E02, ADR-0015).

v1 обмежується однією главою (за замовчуванням ChapterNo="01") --
тестова підмножина, повне дерево 17000+ вузлів -- окреме майбутнє
рішення (ADR-0015 п.4). Файл сам по собі НЕ зберігається в git
(розмір, дублювання зовнішнього джерела) -- шлях передається через
WHO_ICD11_DATA_DIR, той самий принцип, що MEDIC_DATA_DIR.

Побудова дерева -- двопрохідна, бо файл не містить прямого parent_id:
  Прохід 1: {BlockId: Linearization URI} для ВСІХ рядків файлу
            (не тільки обраної глави -- предок з іншим ChapterNo
            теоретично можливий, безпечніше не звужувати заздалегідь).
  Прохід 2: для кожного рядка обраної глави -- останній непорожній
            серед Grouping5..Grouping1 (від найглибшого до
            найвищого) -- це BlockId найближчого предка. Немає
            жодного заповненого Grouping -- це рядок глави
            (Chapter), parent_id = None.

Title у WHO-файлі має провідний текстовий маркер вкладеності
("- ", "- - ", ...), що дублює наше структурне node_kind/parent_id --
прибирається окремо (_strip_depth_marker).

ukrainian_title завжди None після цього importer-а -- джерело
англомовне, translation_status=MISSING для всіх імпортованих вузлів
(не AUTO_IMPORTED: жодного перекладу не було спроби зробити, це
структурний skeleton-імпорт). Переклад -- окремий майбутній крок.
"""
import os
import re
from pathlib import Path

from openpyxl import load_workbook

from app.modules.icd11.domain.entities import (
    ICD11Node,
    NodeKind,
    SpecialCode,
    TranslationStatus,
)

DEFAULT_ENV_VAR = "WHO_ICD11_DATA_DIR"
DEFAULT_FILENAME = "SimpleTabulation-ICD-11-MMS-en.xlsx"
FALLBACK_SOURCE_RELEASE = "2024-01"

GROUPING_COLUMNS = ["Grouping5", "Grouping4", "Grouping3", "Grouping2", "Grouping1"]


def resolve_who_data_dir(explicit_path: str | None = None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
    else:
        env_value = os.environ.get(DEFAULT_ENV_VAR)
        if not env_value:
            raise RuntimeError(
                f"Шлях до WHO ICD-11 файлу не заданий. Встанови змінну "
                f'середовища {DEFAULT_ENV_VAR} (наприклад: '
                f'$env:{DEFAULT_ENV_VAR}="D:\\FHOS\\external_data\\icd11") '
                f"або передай explicit_path напряму."
            )
        path = Path(env_value)

    if not path.is_dir():
        raise FileNotFoundError(f"Тека з WHO ICD-11 даними не знайдена: {path}")

    return path


def _strip_depth_marker(title: str) -> str:
    return re.sub(r"^(?:-\s*)+", "", title or "").strip()


def _parse_special_code(code: str | None) -> SpecialCode:
    if not code:
        return SpecialCode.NONE
    normalized = code.strip().upper()
    if normalized.endswith(".Y"):
        return SpecialCode.Y
    if normalized.endswith(".Z"):
        return SpecialCode.Z
    return SpecialCode.NONE


def _find_version_column(header: list) -> str | None:
    for col in header:
        if col and str(col).strip().lower().startswith("version"):
            return col
    return None


def _normalize_chapter(value: str) -> str:
    return value.zfill(2) if value.isdigit() else value


def load_icd11_chapter_from_xlsx(
    who_data_dir: str | None = None,
    chapter_no: str = "01",
    filename: str = DEFAULT_FILENAME,
) -> list[ICD11Node]:
    data_dir = resolve_who_data_dir(who_data_dir)
    xlsx_path = data_dir / filename
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Файл не знайдено: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if len(wb.sheetnames) > 1:
        print(
            f"УВАГА: у файлі кілька листів: {wb.sheetnames} -- перевір, "
            f"чи не містить один із них ліцензійний текст"
        )
    sheet = wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    header = list(next(rows_iter))
    print(f"Заголовок файлу ({len(header)} колонок): {header}")

    version_column = _find_version_column(header)
    if version_column:
        print(f"Знайдено колонку версії: {version_column!r}")
    else:
        print(f"Колонку версії не знайдено -- fallback {FALLBACK_SOURCE_RELEASE!r}")

    raw_rows = [dict(zip(header, values)) for values in rows_iter]
    print(f"Прочитано рядків даних (без заголовка): {len(raw_rows)}")

    # Прохід 1: BlockId -> Linearization URI, ПО ВСЬОМУ файлу (не тільки
    # обраній главі -- безпечніше не звужувати джерело мапи заздалегідь).
    blockid_to_uri: dict[str, str] = {}
    for row in raw_rows:
        block_id = row.get("BlockId")
        uri = row.get("Linearization URI")
        if block_id and uri:
            blockid_to_uri[str(block_id).strip()] = str(uri).strip()
    print(f"Побудовано BlockId -> URI мапу: {len(blockid_to_uri)} записів")

    normalized_filter = _normalize_chapter(str(chapter_no).strip())

    nodes: list[ICD11Node] = []
    skipped_wrong_chapter = 0
    skipped_unknown_class_kind = 0
    skipped_invalid = 0
    unresolved_parents = 0

    for row in raw_rows:
        row_chapter_raw = row.get("ChapterNo")
        row_chapter_str = str(row_chapter_raw).strip() if row_chapter_raw is not None else ""
        if _normalize_chapter(row_chapter_str) != normalized_filter:
            skipped_wrong_chapter += 1
            continue

        class_kind_raw = (row.get("ClassKind") or "").strip().lower()
        try:
            node_kind = NodeKind(class_kind_raw)
        except ValueError:
            skipped_unknown_class_kind += 1
            continue

        parent_block_id = None
        for grouping_col in GROUPING_COLUMNS:
            value = row.get(grouping_col)
            if value:
                parent_block_id = str(value).strip()
                break

        parent_id = None
        if parent_block_id:
            parent_id = blockid_to_uri.get(parent_block_id)
            if parent_id is None:
                unresolved_parents += 1

        uri = row.get("Linearization URI")
        title = row.get("Title") or ""
        code = row.get("Code")
        source_release = (
            str(row.get(version_column)).strip()
            if version_column and row.get(version_column)
            else FALLBACK_SOURCE_RELEASE
        )

        try:
            node = ICD11Node(
                id=str(uri).strip() if uri else "",
                parent_id=parent_id,
                icd_code=str(code).strip() if code else None,
                english_title=_strip_depth_marker(str(title)),
                ukrainian_title=None,
                translation_status=TranslationStatus.MISSING,
                node_kind=node_kind,
                special_code=_parse_special_code(str(code) if code else None),
                sort_order=len(nodes) + 1,
                source_release=source_release,
            )
        except ValueError:
            skipped_invalid += 1
            continue

        nodes.append(node)

    print(f"Розділ {chapter_no}: імпортовано вузлів: {len(nodes)}")
    print(f"Пропущено (інша глава): {skipped_wrong_chapter}")
    print(f"Пропущено (невідомий ClassKind): {skipped_unknown_class_kind}")
    print(f"Пропущено (невалідний вузол, напр. порожній id): {skipped_invalid}")
    print(f"Нерозв'язаних parent у межах глави (BlockId без відповідного URI): {unresolved_parents}")

    return nodes
